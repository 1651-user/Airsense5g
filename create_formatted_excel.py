"""
Create Excel with properly formatted timestamp column
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

print("Creating Excel with formatted timestamp...")

# Read the data
df = pd.read_excel('combined_sensor_data_20251222_132447.xlsx')

# Rename for clarity
df.rename(columns={'received_at': 'Timestamp'}, inplace=True)

# Reorder columns
cols = ['Timestamp'] + [c for c in df.columns if c != 'Timestamp']
df = df[cols]

# Save to Excel
output_file = 'sensor_data_formatted.xlsx'
df.to_excel(output_file, index=False)

# Format the Excel file
wb = load_workbook(output_file)
ws = wb.active

# Set column widths
ws.column_dimensions['A'].width = 25  # Timestamp column
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
    ws.column_dimensions[col].width = 12

# Format timestamp column as datetime
for row in range(2, len(df) + 2):  # Skip header
    cell = ws[f'A{row}']
    cell.number_format = 'YYYY-MM-DD HH:MM:SS'

# Save
wb.save(output_file)

print(f"Created: {output_file}")
print(f"  Rows: {len(df)}")
print(f"  Timestamp column width: 25")
print(f"  Format: YYYY-MM-DD HH:MM:SS")

# Open the file
import os
os.system(f'start {output_file}')
